# 1, import from installed library
import cv2
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment

# 2, import from source code
import src._common as _com

# 3, import from python library
import os
import glob

###############################################################


def resize_image(max_size):

    img = cv2.imread(_com.var.common_filename + _com.var.filename_ext)

    height, width = img.shape[:2]

    if (max(width, height) > max_size):

        aspect = width / height

        if width > height:
            nw = max_size
            nh = round(nw / aspect)
        else:
            nh = max_size
            nw = round(nh * aspect)

        dst = cv2.resize(img, dsize=(nw, nh), interpolation = cv2.INTER_LINEAR)
        cv2.imwrite(_com.var.common_filename + "_resized" + _com.var.filename_ext  , dst)

        print(f"file is resized from {_com.var.common_filename} to {_com.var.common_filename}_resized")

        _com.var.common_filename = _com.var.common_filename + "_resized"


###############################################################
def generate_xlsx():
    filelist = generate_filelist()

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]  # target is first worksheet
    ws.title = 'compare'   # set sheet tile

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 90.71
    ws.column_dimensions['C'].width = 90.71
    ws.column_dimensions['D'].width = 90.71
    ws.column_dimensions['E'].width = 90.71

    target_img = [_com.var.filename_ext,
                  '_sam_withMask.png',
                  '_sam_Temperature_boxplot.png'
                  ]
    target_img_info = ['Original image',
                       'SAM',
                       'Boxplot']

    ws.cell(row = 1 , column = 1, value = 'Filename')
    for j, target in enumerate(target_img_info):
        ws.cell(row=1 , column=j + 2, value=target)

    # length = len(filelist)

    for i, file in enumerate(filelist):
        # print(str(i) + '/' + str(length))

        ws.row_dimensions[i + 2].height = 360

        ws.cell(row = i + 2, column = 1, value = file).font = Font(size = 16)
        ws.cell(row = i + 2, column = 1, value = file).alignment = Alignment(vertical = "center")

        for j, target in enumerate(target_img):
            img = openpyxl.drawing.image.Image(file + target)
            cell_address = ws.cell(row = i + 2, column = j + 2).coordinate
            img.anchor = cell_address
            ws.add_image(img)

        # if (i > 2):
        #    break

    # save file
    wb.save('compare.xlsx')

###############################################################


def check_require_files() -> bool:

    if _com.var.use_temp is False:
        return True

    csv_filename = _com.var.common_filename + '.csv'
    xlsx_filename = _com.var.common_filename + '.xlsx'

    # require csv or xls
    if (os.path.isfile(csv_filename) or os.path.isfile(xlsx_filename)):
        return True
    else:
        print(f"{xlsx_filename} is not found in the directory")
        return False


###############################################################


def generate_filelist() -> list:
    """_summary_

    path以下にある**_Visual_IR.jpeg画像のリストを返す

    Args:
        path (_type_): 画像があるディレクトリ

    Returns:
        filelist: common_filename(**.jpegのうち**の部分)が格納されている配列
    """

    filelist = []

    if (os.path.isfile(_com.var.path)):
        name = _com.var.path
        if _com.var.filename_ext in name:
            print(f"file: {name}")

            jpeg_filename = name
            common_filename = jpeg_filename.replace(_com.var.filename_ext, "")

            filelist.append(common_filename)
    else:

        for name in glob.glob(os.path.join(_com.var.path, '*')):
            if _com.var.filename_ext in name:
                print(f"file: {name}")

                jpeg_filename = os.path.basename(name)
                common_filename = jpeg_filename.replace(_com.var.filename_ext, "")
                common_filename = os.path.join(_com.var.path, common_filename)

                filelist.append(common_filename)

    filelist = list(set(filelist))

    filelist = sorted(filelist)

    return filelist
